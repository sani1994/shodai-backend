import math
import uuid
from datetime import datetime, timedelta
from random import randint

from decouple import config
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.core.validators import validate_email
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils import timezone
from django.utils.crypto import get_random_string
from django_q.tasks import async_task
from num2words import num2words
from openpyxl import Workbook
from rest_framework import status
from rest_framework.generics import get_object_or_404
from rest_framework.views import APIView
from base.views import CustomPageNumberPagination, field_validation, type_validation, coupon_checker
from coupon.models import CouponCode, CouponSettings, CouponUser, CouponUsageHistory
from offer.models import Offer, CartOffer
from order.models import Order, InvoiceInfo, OrderProduct, DeliveryCharge, TimeSlot, DiscountInfo, PreOrderSetting, \
    PreOrder
from producer.models import ProducerProductRequest
from product.models import Product, ProductMeta
from shodai_admin.serializers import AdminUserProfileSerializer, OrderListSerializer, OrderDetailSerializer, \
    ProductSearchSerializer, TimeSlotSerializer, CustomerSerializer, DeliveryChargeOfferSerializer, \
    UserProfileSerializer, ProductMetaSerializer, order_status_all, PreOrderSettingListSerializer, \
    PreOrderSettingDetailSerializer, ProducerProductSerializer, PreOrderListSerializer, PreOrderDetailSerializer, DeliveryZoneSerializer
from shodai.permissions import IsAdminUserQP
from user.models import UserProfile, Address

from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response

from utility.models import DeliveryZone
from utility.notification import send_sms
from utility.pdf import render_to_pdf

all_order_status = {
    'Ordered': 'OD',
    'Order Accepted': 'OA',
    'Order Ready': 'RE',
    'Order at Delivery': 'OAD',
    'Order Completed': 'COM',
    'Order Cancelled': 'CN'
}


class AdminUserProfile(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        serializer = AdminUserProfileSerializer(request.user)
        if serializer:
            return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AdminUserRegistration(APIView):

    def post(self, request):
        data = request.data
        required_fields = ['name',
                           'mobile_number',
                           'email',
                           'password']
        is_valid = field_validation(required_fields, data)

        if is_valid:
            string_fields = [data['name'],
                             data['mobile_number'],
                             data['email'],
                             data['password']]
            is_valid = type_validation(string_fields, str)
        if is_valid and len(data["mobile_number"]) == 14 and \
                data["mobile_number"].startswith('+8801') and data["mobile_number"][1:].isdigit():
            pass
        else:
            is_valid = False
        if is_valid and data["email"]:
            try:
                validate_email(data["email"])
            except Exception:
                is_valid = False
        if not is_valid or not data['name'] or not data['password']:
            return Response({
                "status": "failed",
                "message": "Invalid request!"
            }, status=status.HTTP_400_BAD_REQUEST)
        else:
            try:
                mobile_user_instance = UserProfile.objects.get(mobile_number=data["mobile_number"])
            except UserProfile.DoesNotExist:
                mobile_user_instance = None

            try:
                email_user_instance = UserProfile.objects.get(email=data["email"])
            except UserProfile.DoesNotExist:
                email_user_instance = None

            if not mobile_user_instance and not email_user_instance:
                user_instance = UserProfile.objects.create(username=data["mobile_number"],
                                                           first_name=data["name"],
                                                           last_name="",
                                                           email=data["email"],
                                                           mobile_number=data["mobile_number"],
                                                           user_type="CM",
                                                           created_on=timezone.now(),
                                                           verification_code=randint(100000, 999999),
                                                           is_approved=True)
                user_instance.set_password(data['password'])
                user_instance.save()
                return Response({
                    "status": "success",
                    "message": "User Created."}, status=status.HTTP_200_OK)
            else:
                return Response({
                    "status": "failed",
                    "message": "Mobile or Email Already Exists!"
                }, status=status.HTTP_400_BAD_REQUEST)


class Login(APIView):

    def post(self, request):
        data = request.data
        if 'username' not in data or 'password' not in data:
            return Response({
                "status": "failed",
                "message": "Invalid request!"
            }, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(username=data['username'], password=data['password'])

        if not user or not user.is_staff:
            return Response({
                "status": "failed",
                "message": "Invalid username or password!"
            }, status=status.HTTP_401_UNAUTHORIZED)
        else:
            token, _ = Token.objects.get_or_create(user=user)
            return Response({
                "message": "Login successful.",
                "status": "success",
                "username": user.username,
                'access_token': token.key,
            }, status=status.HTTP_200_OK)


class Logout(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        Token.objects.get(key=request.auth).delete()
        return Response({
            "status": "success",
            "message": "Logout successful."
        }, status=status.HTTP_200_OK)


class TokenViewAPITest(APIView):  # Sample API test with Authentication
    permission_classes = [IsAdminUser]

    def get(self, request):
        if not request.user.has_perms(['authtoken.view_token']):
            return Response({
                "status": "failed",
                "message": "You are not authorized to do this action!"
            }, status=status.HTTP_401_UNAUTHORIZED)

        token = Token.objects.get(user=request.user)
        data = {'token': token.key}
        return Response(data, status=status.HTTP_200_OK)


class OrderList(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        search = request.query_params.get('search')
        sort_by = request.query_params.get('sort_by', '')
        sort_type = request.query_params.get('sort_type', 'dsc')
        date_type = request.query_params.get('date_type', 'placed_on')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        order_status = all_order_status.get(request.query_params.get('order_status'))
        if not getattr(Order, sort_by, False):
            sort_by = 'placed_on'
        if sort_type != 'asc':
            sort_by = '-' + sort_by
        try:
            date_from = timezone.make_aware(datetime.strptime(date_from, "%Y-%m-%d"))
        except Exception:
            if date_type == 'placed_on':
                date_from = Order.objects.first().placed_on
            else:
                date_from = Order.objects.first().delivery_date_time
        try:
            date_to = timezone.make_aware(datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except Exception:
            date_to = timezone.now()

        if date_type == 'placed_on':
            if search and order_status:
                if search.startswith("01") and len(search) == 11:
                    queryset = Order.objects.filter(order_status=order_status,
                                                    user__mobile_number='+88' + search,
                                                    placed_on__gte=date_from,
                                                    placed_on__lt=date_to).order_by(sort_by)
                else:
                    queryset = Order.objects.filter(order_status=order_status,
                                                    order_number__icontains=search,
                                                    placed_on__gte=date_from,
                                                    placed_on__lt=date_to).order_by(sort_by)

            elif search and not order_status:
                if search.startswith("01") and len(search) == 11:
                    queryset = Order.objects.filter(user__mobile_number='+88' + search,
                                                    placed_on__gte=date_from,
                                                    placed_on__lt=date_to).order_by(sort_by)
                else:
                    queryset = Order.objects.filter(order_number__icontains=search,
                                                    placed_on__gte=date_from,
                                                    placed_on__lt=date_to).order_by(sort_by)

            elif not search and order_status:
                queryset = Order.objects.filter(order_status=order_status,
                                                placed_on__gte=date_from,
                                                placed_on__lt=date_to).order_by(sort_by)
            else:
                queryset = Order.objects.filter(placed_on__gte=date_from,
                                                placed_on__lt=date_to).order_by(sort_by)
        else:
            if search and order_status:
                if search.startswith("01") and len(search) == 11:
                    queryset = Order.objects.filter(order_status=order_status,
                                                    user__mobile_number='+88' + search,
                                                    delivery_date_time__gte=date_from,
                                                    delivery_date_time__lt=date_to).order_by(sort_by)
                else:
                    queryset = Order.objects.filter(order_status=order_status,
                                                    order_number__icontains=search,
                                                    delivery_date_time__gte=date_from,
                                                    delivery_date_time__lt=date_to).order_by(sort_by)

            elif search and not order_status:
                if search.startswith("01") and len(search) == 11:
                    queryset = Order.objects.filter(user__mobile_number='+88' + search,
                                                    delivery_date_time__gte=date_from,
                                                    delivery_date_time__lt=date_to).order_by(sort_by)
                else:
                    queryset = Order.objects.filter(order_number__icontains=search,
                                                    delivery_date_time__gte=date_from,
                                                    delivery_date_time__lt=date_to).order_by(sort_by)

            elif not search and order_status:
                queryset = Order.objects.filter(order_status=order_status,
                                                delivery_date_time__gte=date_from,
                                                delivery_date_time__lt=date_to).order_by(sort_by)
            else:
                queryset = Order.objects.filter(delivery_date_time__gte=date_from,
                                                delivery_date_time__lt=date_to).order_by(sort_by)
        paginator = CustomPageNumberPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = OrderListSerializer(result_page, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)


class OrderDetail(APIView):
    permission_classes = [IsAdminUser]
    """
    Get, update order
    """

    def get(self, request, id):
        order = get_object_or_404(Order, id=id)
        serializer = OrderDetailSerializer(order)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, id):
        data = request.data
        offer_id = data.get('offer_id')
        products = data['products']
        required_fields = ['delivery_date',
                           'delivery_time_slot_id',
                           'delivery_address',
                           'contact_number',
                           'order_status',
                           'products',
                           'note',
                           'coupon_code',
                           'additional_discount',
                           'delivery_zone_id']
        is_valid = field_validation(required_fields, data)

        if is_valid:
            string_fields = [data['delivery_date'],
                             data['delivery_address'],
                             data['contact_number'],
                             data['order_status'],
                             data['note'],
                             data['coupon_code']]
            is_valid = type_validation(string_fields, str)

        order = Order.objects.filter(id=id).first()
        if is_valid and order:
            all_order_products = OrderProduct.objects.filter(order=order)

            order_products = []
            order_product_list = []
            for item in all_order_products:
                product_data = {'product_id': item.product.id,
                                'product_quantity': item.order_product_qty}
                order_product_list.append(item.product.id)
                order_products.append(product_data)
        else:
            is_valid = False

        if is_valid and isinstance(products, list) and products:
            required_fields = ['product_id', 'product_quantity']

            product_list = []
            for item in products:
                is_valid = field_validation(required_fields, item)
                if is_valid and isinstance(item['product_id'], int):
                    if item['product_id'] not in product_list:
                        product_list.append(item['product_id'])
                        if item['product_id'] in order_product_list:
                            product_exist = Product.objects.filter(id=item['product_id'])
                        else:
                            product_exist = Product.objects.filter(id=item['product_id'], is_approved=True)
                        if not product_exist:
                            is_valid = False
                        if is_valid:
                            decimal_allowed = product_exist[0].decimal_allowed
                            if not decimal_allowed and not isinstance(item['product_quantity'], int) and \
                                    not item['product_quantity'] > 0:
                                is_valid = False
                            elif decimal_allowed and not isinstance(item['product_quantity'], (float, int)) and \
                                    not item['product_quantity'] > 0:
                                is_valid = False
                        if is_valid and decimal_allowed:
                            item['product_quantity'] = math.floor(item['product_quantity'] * 10 ** 3) / 10 ** 3
                    else:
                        is_valid = False
                else:
                    is_valid = False
                if not is_valid:
                    break
        else:
            is_valid = False

        if is_valid and isinstance(data['delivery_time_slot_id'], int):
            time = TimeSlot.objects.filter(id=data['delivery_time_slot_id'])
            if time:
                delivery_date_time = data['delivery_date'] + str(time[0].time)
                try:
                    delivery_date_time = timezone.make_aware(datetime.strptime(delivery_date_time, "%Y-%m-%d%H:%M:%S"))
                except Exception:
                    is_valid = False
            else:
                is_valid = False
        else:
            is_valid = False

        if is_valid and len(data["contact_number"]) == 14 and \
                data["contact_number"].startswith('+8801') and data["contact_number"][1:].isdigit():
            pass
        else:
            is_valid = False

        if is_valid and isinstance(data['delivery_zone_id'], int) and data['delivery_zone_id']:
            zone = DeliveryZone.objects.filter(id=data['delivery_zone_id'], is_approved=True).first()
        else:
            is_valid = False

        if not is_valid or not data['delivery_address'] or data['order_status'] not in all_order_status or \
                not isinstance(data['additional_discount'], (float, int)) or not zone:
            return Response({
                "status": "failed",
                "message": "Invalid request!"
            }, status=status.HTTP_400_BAD_REQUEST)
        else:
            products_updated = True
            if order.order_status == 'COM' or order.order_status == 'CN' or \
                    data['order_status'] == 'Order Completed' or data['order_status'] == 'Order Cancelled':
                products_updated = False

            if products_updated and len(order_products) == len(products):
                for i in products:
                    if i not in order_products:
                        break
                else:
                    products_updated = False

            data["delivery_address"] = data["delivery_address"][:500]
            if not order.address or data["delivery_address"] != order.address.road:
                delivery_address = Address.objects.filter(road=data["delivery_address"]).first()
                if not delivery_address:
                    delivery_address = Address.objects.create(road=data["delivery_address"],
                                                              city="Dhaka",
                                                              district="Dhaka",
                                                              country="Bangladesh",
                                                              zip_code="",
                                                              user=order.user)
            else:
                delivery_address = order.address

            invoice = InvoiceInfo.objects.filter(order_number=order).order_by('-created_on').first()
            delivery_charge = invoice.delivery_charge

            is_delivery_discount = DiscountInfo.objects.filter(discount_type='DC', invoice=invoice)
            delivery_charge_discount = is_delivery_discount[0].discount_amount if is_delivery_discount else 0
            if offer_id and isinstance(offer_id, int):
                today = timezone.now()
                offer = Offer.objects.filter(id=offer_id, offer_types="DD", is_approved=True,
                                             offer_starts_in__lte=today, offer_ends_in__gte=today)
                if offer:
                    cart_offer = CartOffer.objects.get(offer=offer[0])
                    if cart_offer:
                        delivery_charge_without_offer = DeliveryCharge.objects.get().delivery_charge_inside_dhaka
                        delivery_charge = cart_offer.updated_delivery_charge
                        delivery_charge_discount = delivery_charge_without_offer - delivery_charge

            coupon_discount = prev_additional_discount = 0
            coupon_discount_description = ""
            is_coupon_discount = DiscountInfo.objects.filter(discount_type='CP', invoice=invoice)
            if is_coupon_discount:
                coupon_discount = is_coupon_discount[0].discount_amount
                coupon_discount_description = is_coupon_discount[0].discount_description
            elif data['coupon_code']:
                discount_amount, coupon, is_using, _ = coupon_checker(data['coupon_code'], products, order.user)
                if discount_amount:
                    coupon_discount = discount_amount
                    coupon_discount_description = 'Coupon Code: {}'.format(data['coupon_code'])
                    is_using.remaining_usage_count -= 1
                    is_using.save()
                    coupon.max_usage_count -= 1
                    coupon.save()

            additional_discount = data['additional_discount']
            is_additional_discount = DiscountInfo.objects.filter(discount_type='AD', invoice=invoice)
            if is_additional_discount:
                prev_additional_discount = is_additional_discount[0].discount_amount

            if products_updated:
                if "-" in order.order_number:
                    x = order.order_number.split("-")
                    x[1] = int(x[1]) + 1
                    order_number = x[0] + "-" + str(x[1])
                else:
                    order_number = order.order_number + "-1"

                is_used = Order.objects.filter(order_number=order_number).count()
                if is_used:
                    return Response({
                        "status": "failed",
                        "message": "Invalid request!"
                    }, status=status.HTTP_400_BAD_REQUEST)

                existing_order = order.id
                order.order_status = 'CN'
                order.save()

                order = Order.objects.create(user=order.user,
                                             placed_on=order.placed_on,
                                             platform=order.platform,
                                             order_number=order_number,
                                             delivery_date_time=delivery_date_time,
                                             delivery_place=order.delivery_place,
                                             delivery_zone=zone,
                                             lat=order.lat,
                                             long=order.long,
                                             order_status="OA",
                                             address=delivery_address,
                                             contact_number=data['contact_number'],
                                             created_by=request.user)

                total_vat = total = total_price = total_op_price = total_price_regular_product = 0
                for p in products:
                    product = Product.objects.get(id=p["product_id"])
                    is_op_unchanged = OrderProduct.objects.filter(product=product,
                                                                  order=existing_order)
                    if is_op_unchanged and is_op_unchanged[0].order_product_qty >= p["product_quantity"]:
                        op = OrderProduct.objects.create(product=product,
                                                         order=order,
                                                         order_product_price=is_op_unchanged[0].order_product_price,
                                                         product_price=is_op_unchanged[0].product_price,
                                                         order_product_qty=p["product_quantity"])
                    else:
                        op = OrderProduct.objects.create(product=product,
                                                         order=order,
                                                         order_product_qty=p["product_quantity"])
                    total_price += float(op.product_price) * op.order_product_qty
                    total_op_price += op.order_product_price * op.order_product_qty
                    total += float(op.order_product_price_with_vat) * op.order_product_qty
                    total_vat += float(op.order_product_price_with_vat - op.order_product_price) * op.order_product_qty
                    if op.product_price == op.order_product_price:
                        total_price_regular_product += float(op.product_price) * op.order_product_qty

                if is_coupon_discount:
                    discount_amount, _, _, _ = coupon_checker(is_coupon_discount[0].coupon.coupon_code,
                                                              products, order.user, is_used=True)
                    if discount_amount:
                        coupon_discount = discount_amount
                        coupon_discount_description = is_coupon_discount[0].discount_description

                order.order_total_price = round(total + delivery_charge - coupon_discount - additional_discount)
                order.total_vat = total_vat
                order.note = data['note'][:500]
            else:
                order.order_total_price = round(order.order_total_price - invoice.delivery_charge + delivery_charge +
                                                prev_additional_discount - additional_discount)
                order.invoice_number = "SHD" + str(uuid.uuid4())[:8].upper()
                order.delivery_date_time = delivery_date_time
                order.contact_number = data['contact_number']
                order.order_status = all_order_status[data['order_status']]
                order.address = delivery_address
                order.note = data['note'][:500]
                order.modified_by = request.user
                order.delivery_zone = zone

                if not is_coupon_discount and data['coupon_code']:
                    order.order_total_price -= coupon_discount

            if order.order_total_price < 0:
                order.order_total_price += additional_discount
                additional_discount = 0

            if products_updated:
                product_discount = total_price - total_op_price
            else:
                is_product_discount = DiscountInfo.objects.filter(discount_type='PD', invoice=invoice)
                product_discount = is_product_discount[0].discount_amount if is_product_discount else 0

            discount_amount = delivery_charge_discount + coupon_discount + product_discount + additional_discount

            if order.user.first_name and order.user.last_name:
                billing_person_name = order.user.first_name + " " + order.user.last_name
            elif order.user.first_name:
                billing_person_name = order.user.first_name
            else:
                billing_person_name = ""
            payment_method = 'CASH_ON_DELIVERY' if products_updated else invoice.payment_method
            paid_status = False if products_updated else invoice.paid_status
            paid_on = None if products_updated else invoice.paid_on
            new_invoice = InvoiceInfo.objects.create(invoice_number=order.invoice_number,
                                                     billing_person_name=billing_person_name,
                                                     billing_person_email=order.user.email,
                                                     billing_person_mobile_number=order.user.mobile_number,
                                                     delivery_contact_number=order.contact_number,
                                                     delivery_address=delivery_address.road,
                                                     delivery_date_time=order.delivery_date_time,
                                                     delivery_charge=delivery_charge,
                                                     discount_amount=discount_amount,
                                                     net_payable_amount=order.order_total_price,
                                                     payment_method=payment_method,
                                                     paid_status=paid_status,
                                                     paid_on=paid_on,
                                                     order_number=order,
                                                     user=order.user,
                                                     created_by=request.user)
            order.save()

            if coupon_discount:
                DiscountInfo.objects.create(discount_amount=coupon_discount,
                                            discount_type='CP',
                                            discount_description=coupon_discount_description,
                                            coupon=is_coupon_discount[0].coupon if is_coupon_discount else coupon,
                                            invoice=new_invoice)

            if product_discount:
                DiscountInfo.objects.create(discount_amount=product_discount,
                                            discount_type='PD',
                                            discount_description='Product Offer Discount',
                                            invoice=new_invoice)

            if delivery_charge_discount:
                DiscountInfo.objects.create(discount_amount=delivery_charge_discount,
                                            discount_type='DC',
                                            discount_description='Offer ID: {}'.format(offer_id)
                                            if offer_id else is_delivery_discount[0].discount_description,
                                            offer=offer[0] if offer_id else is_delivery_discount[0].offer,
                                            invoice=new_invoice)

            if additional_discount:
                DiscountInfo.objects.create(discount_amount=additional_discount,
                                            discount_type='AD',
                                            discount_description='Additional Discount',
                                            invoice=new_invoice)
            return Response({
                "status": "success",
                "message": "Order updated.",
                "order_id": order.id}, status=status.HTTP_200_OK)


class CreateOrder(APIView):
    permission_classes = [IsAdminUser]
    """
    Create order
    """

    def post(self, request):
        data = request.data
        offer_id = data.get('offer_id', None)
        required_fields = ['delivery_date',
                           'delivery_time_slot_id',
                           'delivery_address',
                           'contact_number',
                           'customer',
                           'products',
                           'note',
                           'coupon_code',
                           'additional_discount',
                           'delivery_zone_id']
        is_valid = field_validation(required_fields, data)

        if is_valid:
            customer = data['customer']
            products = data['products']

            string_fields = [data['delivery_date'],
                             data['delivery_address'],
                             data['contact_number'],
                             data['note'],
                             data['coupon_code']]
            is_valid = type_validation(string_fields, str)

        if is_valid:
            required_fields = ['name', 'mobile_number', 'email']
            is_valid = field_validation(required_fields, customer)
            if is_valid:
                string_fields = [customer["name"],
                                 customer["mobile_number"],
                                 customer["email"]]
                is_valid = type_validation(string_fields, str)
            if is_valid and len(customer["mobile_number"]) == 14 and \
                    customer["mobile_number"].startswith('+8801') and customer["mobile_number"][1:].isdigit():
                pass
            else:
                is_valid = False
            if is_valid and customer["email"]:
                try:
                    validate_email(customer["email"])
                except Exception:
                    is_valid = False

        if is_valid and isinstance(products, list) and products:
            required_fields = ['product_id', 'product_quantity']
            product_list = []
            for item in products:
                is_valid = field_validation(required_fields, item)
                if is_valid and isinstance(item['product_id'], int):
                    if item['product_id'] not in product_list:
                        product_list.append(item['product_id'])
                        product_exist = Product.objects.filter(id=item['product_id'], is_approved=True)
                        if not product_exist:
                            is_valid = False
                        if is_valid:
                            decimal_allowed = product_exist[0].decimal_allowed
                            if not decimal_allowed and not isinstance(item['product_quantity'], int) and \
                                    not item['product_quantity'] > 0:
                                is_valid = False
                            elif decimal_allowed and not isinstance(item['product_quantity'], (float, int)) and \
                                    not item['product_quantity'] > 0:
                                is_valid = False
                        if is_valid and decimal_allowed:
                            item['product_quantity'] = math.floor(item['product_quantity'] * 10 ** 3) / 10 ** 3
                    else:
                        is_valid = False
                else:
                    is_valid = False
                if not is_valid:
                    break
        else:
            is_valid = False

        if is_valid and isinstance(data['delivery_time_slot_id'], int):
            time = TimeSlot.objects.filter(id=data['delivery_time_slot_id'])
            if time:
                delivery_date_time = data['delivery_date'] + str(time[0].time)
                try:
                    delivery_date_time = timezone.make_aware(datetime.strptime(delivery_date_time, "%Y-%m-%d%H:%M:%S"))
                except Exception:
                    is_valid = False
            else:
                is_valid = False
        else:
            is_valid = False

        if is_valid and len(data["contact_number"]) == 14 and \
                data["contact_number"].startswith('+8801') and data["contact_number"][1:].isdigit():
            pass
        else:
            is_valid = False
            
        if is_valid and isinstance(data['delivery_zone_id'], int) and data['delivery_zone_id']:
            zone = DeliveryZone.objects.filter(id=data['delivery_zone_id'], is_approved=True).first()
        else:
            is_valid = False

        if not is_valid or not data['delivery_address'] or not isinstance(data['additional_discount'], (float, int)) \
                or not zone:
            return Response({
                "status": "failed",
                "message": "Invalid request!"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            is_new_user = False
            user_instance = UserProfile.objects.get(mobile_number=customer["mobile_number"])
        except UserProfile.DoesNotExist:
            is_new_user = True
            user_instance = None

        delivery_charge = DeliveryCharge.objects.get().delivery_charge_inside_dhaka
        delivery_charge_discount = 0
        if offer_id and isinstance(offer_id, int):
            today = timezone.now()
            offer = Offer.objects.filter(id=offer_id, offer_types="DD", is_approved=True,
                                         offer_starts_in__lte=today, offer_ends_in__gte=today)
            if offer:
                cart_offer = CartOffer.objects.get(offer=offer[0])
                if cart_offer:
                    delivery_charge_discount = delivery_charge - cart_offer.updated_delivery_charge
                    delivery_charge = cart_offer.updated_delivery_charge

        coupon_discount = 0
        if data['coupon_code'] and user_instance:
            discount_amount, coupon, is_using, _ = coupon_checker(data['coupon_code'], products, user_instance)
            if discount_amount:
                coupon_discount = discount_amount
                is_using.remaining_usage_count -= 1
                is_using.save()
                coupon.max_usage_count -= 1
                coupon.save()

        additional_discount = data['additional_discount']

        if not user_instance:
            user_instance = UserProfile.objects.create(username=customer["mobile_number"],
                                                       first_name=customer["name"],
                                                       last_name="",
                                                       email=customer["email"],
                                                       mobile_number=customer["mobile_number"],
                                                       user_type="CM",
                                                       created_on=timezone.now(),
                                                       verification_code=randint(100000, 999999),
                                                       code_valid_till=timezone.now() + timedelta(minutes=5),
                                                       is_approved=True)
            temp_password = get_random_string(length=6)
            user_instance.set_password(temp_password)
            user_instance.save()

            gift_discount_settings = CouponSettings.objects.get(coupon_type='GC1')
            if gift_discount_settings.is_active:
                gift_coupon = CouponCode.objects.create(coupon_code=str(uuid.uuid4())[:6].upper(),
                                                        name="Sign Up Coupon",
                                                        discount_percent=gift_discount_settings.discount_percent,
                                                        discount_amount=gift_discount_settings.discount_amount,
                                                        max_usage_count=gift_discount_settings.max_usage_count,
                                                        minimum_purchase_limit=gift_discount_settings.minimum_purchase_limit,
                                                        discount_amount_limit=gift_discount_settings.discount_amount_limit,
                                                        expiry_date=timezone.now() + timedelta(
                                                            days=gift_discount_settings.validity_period),
                                                        discount_type=gift_discount_settings.discount_type,
                                                        coupon_code_type='GC1',
                                                        created_by=user_instance)
                CouponUser.objects.create(coupon_code=gift_coupon,
                                          created_for=user_instance,
                                          remaining_usage_count=1,
                                          created_by=user_instance)

            if not settings.DEBUG:
                sms_body = "Dear Customer,\n" + \
                           "We have created an account with temporary password " + \
                           "[{}] based on your order request. ".format(temp_password) + \
                           "Please change your password after login.\n\n" + \
                           "www.shod.ai"
                async_task('utility.notification.send_sms', user_instance.mobile_number, sms_body)

                if gift_discount_settings.is_active:
                    async_task('coupon.tasks.send_coupon_sms',
                               gift_coupon,
                               user_instance.mobile_number)

            if gift_discount_settings.is_active:
                async_task('coupon.tasks.send_coupon_email',
                           gift_coupon,
                           user_instance)

        data["delivery_address"] = data["delivery_address"][:500]
        address = Address.objects.filter(road=data["delivery_address"])
        if not address:
            delivery_address = Address.objects.create(road=data["delivery_address"][:500],
                                                      city="Dhaka",
                                                      district="Dhaka",
                                                      country="Bangladesh",
                                                      zip_code="",
                                                      user=user_instance)
        else:
            delivery_address = address[0]

        order_instance = Order.objects.create(user=user_instance,
                                              platform="AD",
                                              delivery_date_time=delivery_date_time,
                                              delivery_place="Dhaka",
                                              delivery_zone=zone,
                                              lat=23.7733,
                                              long=90.3548,
                                              order_status="OA",
                                              contact_number=data["contact_number"],
                                              address=delivery_address,
                                              created_by=request.user)

        total_vat = total = total_price = total_op_price = 0
        for p in products:
            product = Product.objects.get(id=p["product_id"])
            op = OrderProduct.objects.create(product=product,
                                             order=order_instance,
                                             order_product_qty=p["product_quantity"])
            total_price += float(op.product_price) * op.order_product_qty
            total_op_price += op.order_product_price * op.order_product_qty
            total += float(op.order_product_price_with_vat) * op.order_product_qty
            total_vat += float(op.order_product_price_with_vat - op.order_product_price) * op.order_product_qty

        order_instance.order_total_price = round(total + delivery_charge - coupon_discount - additional_discount)
        order_instance.total_vat = total_vat
        order_instance.note = data['note'][:500]

        if order_instance.order_total_price < 0:
            order_instance.order_total_price += additional_discount
            additional_discount = 0

        order_instance.save()

        referral_discount_settings = CouponSettings.objects.get(coupon_type='RC')
        if referral_discount_settings.is_active:
            is_referral_code_expired = False
            if not is_new_user:
                referral_coupon = CouponCode.objects.filter(coupon_code_type='RC',
                                                            created_by=user_instance).order_by('-created_on')
                if referral_coupon:
                    referral_coupon = referral_coupon[0]
                    if referral_coupon.expiry_date < timezone.now():
                        is_referral_code_expired = True
                else:
                    is_referral_code_expired = True

            if is_new_user or is_referral_code_expired:
                referral_coupon = CouponCode.objects.create(coupon_code=str(uuid.uuid4())[:6].upper(),
                                                            name="Referral Coupon",
                                                            discount_percent=referral_discount_settings.discount_percent,
                                                            discount_amount=referral_discount_settings.discount_amount,
                                                            max_usage_count=referral_discount_settings.max_usage_count,
                                                            minimum_purchase_limit=referral_discount_settings.minimum_purchase_limit,
                                                            discount_amount_limit=referral_discount_settings.discount_amount_limit,
                                                            expiry_date=timezone.now() + timedelta(
                                                                days=referral_discount_settings.validity_period),
                                                            discount_type=referral_discount_settings.discount_type,
                                                            coupon_code_type='RC',
                                                            created_by=user_instance)

            if referral_coupon.max_usage_count > 0:
                if not settings.DEBUG:
                    async_task('coupon.tasks.send_coupon_sms',
                               referral_coupon,
                               user_instance.mobile_number)
                async_task('coupon.tasks.send_coupon_email',
                           referral_coupon,
                           user_instance)

        product_discount = total_price - total_op_price
        discount_amount = delivery_charge_discount + coupon_discount + product_discount + additional_discount

        billing_person_name = user_instance.first_name + " " + user_instance.last_name
        invoice = InvoiceInfo.objects.create(invoice_number=order_instance.invoice_number,
                                             billing_person_name=billing_person_name,
                                             billing_person_email=user_instance.email,
                                             billing_person_mobile_number=user_instance.mobile_number,
                                             delivery_contact_number=order_instance.contact_number,
                                             delivery_address=delivery_address.road,
                                             delivery_date_time=order_instance.delivery_date_time,
                                             delivery_charge=delivery_charge,
                                             discount_amount=discount_amount,
                                             net_payable_amount=order_instance.order_total_price,
                                             payment_method="CASH_ON_DELIVERY",
                                             order_number=order_instance,
                                             user=user_instance,
                                             created_by=request.user)
        if product_discount:
            DiscountInfo.objects.create(discount_amount=total_price - total_op_price,
                                        discount_type='PD',
                                        discount_description='Product Offer Discount',
                                        invoice=invoice)
        if delivery_charge_discount:
            DiscountInfo.objects.create(discount_amount=delivery_charge_discount,
                                        discount_type='DC',
                                        discount_description='Offer ID: {}'.format(offer_id),
                                        offer=offer[0],
                                        invoice=invoice)

        if coupon_discount:
            DiscountInfo.objects.create(discount_amount=coupon_discount,
                                        discount_type='CP',
                                        discount_description='Coupon Code: {}'.format(data['coupon_code']),
                                        coupon=coupon,
                                        invoice=invoice)
            CouponUsageHistory.objects.create(discount_type=coupon.discount_type,
                                              discount_percent=coupon.discount_percent,
                                              discount_amount=coupon_discount,
                                              coupon_code=coupon.coupon_code,
                                              coupon_user=is_using,
                                              invoice_number=invoice,
                                              created_by=user_instance)
        if additional_discount:
            DiscountInfo.objects.create(discount_amount=additional_discount,
                                        discount_type='AD',
                                        discount_description='Additional Discount',
                                        invoice=invoice)

        return Response({'status': 'success',
                         'message': 'Order placed.',
                         "order_id": order_instance.id}, status=status.HTTP_200_OK)


class ProductSearch(APIView):
    permission_classes = [IsAdminUser]
    """
    Get Product by name
    """

    def get(self, request):
        query = request.query_params.get('query', '')
        product = Product.objects.filter(product_name__icontains=query, is_approved=True)[:20]
        serializer = ProductSearchSerializer(product, many=True)
        return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)


class TimeSlotList(APIView):
    permission_classes = [IsAdminUser]
    """
    Get List of Timeslots
    """

    def get(self, request):
        queryset = TimeSlot.objects.filter(allow=True)
        serializer = TimeSlotSerializer(queryset, many=True)
        if serializer:
            return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class OrderStatusList(APIView):
    permission_classes = [IsAdminUser]
    """
    Get All Order Status
    """

    def get(self, request):
        order_status = [
            'Ordered',
            'Order Accepted',
            'Order Ready',
            'Order at Delivery',
            'Order Completed',
            'Order Cancelled',
        ]
        return Response({'status': 'success', 'data': order_status}, status=status.HTTP_200_OK)


class CustomerSearch(APIView):
    permission_classes = [IsAdminUser]
    """
    Get List of Customers by mobile number
    """

    def get(self, request):
        query = request.query_params.get('query', '')
        queryset = UserProfile.objects.filter(mobile_number__icontains=query,
                                              user_type="CM")
        serializer = CustomerSerializer(queryset, many=True)
        if serializer:
            return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class InvoiceDownloadPDF(APIView):
    permission_classes = [IsAdminUserQP]
    """
    Get PDF of Invoice by order ID
    """

    def get(self, request, id):
        order = get_object_or_404(Order, id=id)
        product_list = OrderProduct.objects.filter(order=order)
        matrix = []
        is_product_offer = False
        sub_total = 0
        for p in product_list:
            if p.order_product_price != p.product_price:
                is_product_offer = True
                total_by_offer = float(p.order_product_price) * p.order_product_qty
                sub_total += total_by_offer
                col = [p.product.product_name, p.product.product_unit, p.product_price,
                       p.order_product_price, p.order_product_qty, total_by_offer]
            else:
                total = float(p.product_price) * p.order_product_qty
                sub_total += total
                col = [p.product.product_name, p.product.product_unit, p.product_price,
                       "--", p.order_product_qty, total]
            matrix.append(col)

        invoice = InvoiceInfo.objects.filter(invoice_number=order.invoice_number)
        if not invoice:
            return HttpResponse("Not found")
        invoice = invoice[0]

        customer_name = invoice.billing_person_name
        paid_status = invoice.paid_status
        is_coupon_discount = DiscountInfo.objects.filter(discount_type='CP', invoice=invoice)
        coupon_discount = is_coupon_discount[0].discount_amount if is_coupon_discount else 0

        is_additional_discount = DiscountInfo.objects.filter(discount_type='AD', invoice=invoice)
        additional_discount = is_additional_discount[0].discount_amount if is_additional_discount else 0

        if invoice.payment_method == "CASH_ON_DELIVERY":
            payment_method = "Cash on Delivery"
        else:
            payment_method = "Online Payment"
        data = {
            'customer_name': customer_name,
            'address': order.address,
            'user_email': order.user.email,
            'user_mobile': order.user.mobile_number,
            'order_number': order.order_number,
            'invoice_number': order.invoice_number,
            'created_on': order.created_on,
            'delivery_date': order.delivery_date_time.date(),
            'order_details': matrix,
            'delivery': invoice.delivery_charge,
            'vat': order.total_vat,
            'sub_total': sub_total,
            'total': order.order_total_price,
            'in_words': num2words(order.order_total_price),
            'payment_method': payment_method if paid_status else 'Cash on Delivery',
            'paid_status': paid_status,
            'is_offer': is_product_offer,
            'coupon_discount': coupon_discount,
            'additional_discount': additional_discount,
            'note': order.note if order.note else None,
            'colspan_value': "4" if is_product_offer else "3",
            'downloaded_on': datetime.now().replace(microsecond=0)
        }
        pdf = render_to_pdf('others/invoice.html', data)
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "invoice_of_order_%s.pdf" % order.order_number
        content = "inline; filename=%s" % filename
        response['Content-Disposition'] = content
        return response


class OrderNotification(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        data = request.data
        if isinstance(data.get('order_id'), int) and data.get('notify_type') in ['placed', 'updated']:
            order_instance = get_object_or_404(Order, id=data['order_id'])
            invoices = InvoiceInfo.objects.filter(order_number=order_instance).order_by('-created_on')
            if data['notify_type'] == 'updated' and len(invoices) > 1 and \
                    invoices[0].net_payable_amount == invoices[1].net_payable_amount:
                pass
            else:
                if data['notify_type'] == 'updated':
                    customer_subject = 'Your shodai order (#' + str(order_instance.order_number) + ') has been updated'
                    admin_subject = 'Order (#' + str(order_instance.order_number) + ') has been updated'
                elif data['notify_type'] == 'placed':
                    customer_subject = 'Your shodai order (#' + str(order_instance.order_number) + ') summary'
                    admin_subject = 'Order (#' + str(order_instance.order_number) + ') has been placed'
                invoice = invoices[0]
                delivery_charge = DeliveryCharge.objects.get().delivery_charge_inside_dhaka
                delivery_charge_discount = delivery_charge - invoice.delivery_charge
                client_name = invoice.billing_person_name
                product_list = OrderProduct.objects.filter(order__pk=data['order_id'])
                matrix = []
                is_offer = False
                sub_total = 0
                for p in product_list:
                    if p.order_product_price != p.product_price:
                        is_offer = True
                        total_by_offer = float(p.order_product_price) * p.order_product_qty
                        sub_total += total_by_offer
                        col = [p.product.product_name, p.product.product_unit, p.product_price,
                               p.order_product_price, p.order_product_qty, total_by_offer]
                    else:
                        total = float(p.product_price) * p.order_product_qty
                        sub_total += total
                        col = [p.product.product_name, p.product.product_unit, p.product_price,
                               "--", p.order_product_qty, total]
                    matrix.append(col)

                time_slot = TimeSlot.objects.get(time=timezone.localtime(order_instance.delivery_date_time).time())

                is_coupon_discount = DiscountInfo.objects.filter(discount_type='CP', invoice=invoice)
                coupon_discount = is_coupon_discount[0].discount_amount if is_coupon_discount else 0

                is_additional_discount = DiscountInfo.objects.filter(discount_type='AD', invoice=invoice)
                additional_discount = is_additional_discount[0].discount_amount if is_additional_discount else 0

                """
                To send notification to customer
                """

                content = {'user_name': client_name,
                           'order_number': order_instance.order_number,
                           'shipping_address': order_instance.address.road + " " + order_instance.address.city,
                           'mobile_no': order_instance.contact_number,
                           'order_date': order_instance.created_on.date(),
                           'delivery_date_time': str(
                               order_instance.delivery_date_time.date()) + " ( " + time_slot.slot + " )",
                           'sub_total': sub_total,
                           'vat': order_instance.total_vat,
                           'delivery_charge': delivery_charge,
                           'total': order_instance.order_total_price,
                           'order_details': matrix,
                           'is_product_discount': is_offer,
                           'coupon_discount': coupon_discount,
                           'additional_discount': additional_discount,
                           'delivery_charge_discount': delivery_charge_discount,
                           'saved_amount': invoice.discount_amount,
                           'note': None,
                           'colspan_value': "4" if is_offer else "3"}

                from_email, to = 'noreply@shod.ai', order_instance.user.email
                html_customer = get_template('email/order_notification_customer.html')
                html_content = html_customer.render(content)
                msg = EmailMultiAlternatives(customer_subject, 'shodai', from_email, [to])
                msg.attach_alternative(html_content, "text/html")
                msg.send()

                """
                To send notification to admin
                """

                if invoice.payment_method == "CASH_ON_DELIVERY":
                    payment_method = "Cash on Delivery"
                else:
                    payment_method = "Online Payment"
                content = {'user_name': client_name,
                           'user_mobile': order_instance.user.mobile_number,
                           'order_number': order_instance.order_number,
                           'platform': "Admin",
                           'shipping_address': order_instance.address.road + " " + order_instance.address.city,
                           'mobile_no': order_instance.contact_number,
                           'order_date': order_instance.created_on.date(),
                           'delivery_date_time': str(
                               order_instance.delivery_date_time.date()) + " ( " + time_slot.slot + " )",
                           'invoice_number': invoice.invoice_number,
                           'payment_method': payment_method,
                           'sub_total': sub_total,
                           'vat': order_instance.total_vat,
                           'delivery_charge': delivery_charge,
                           'total': order_instance.order_total_price,
                           'order_details': matrix,
                           'is_product_discount': is_offer,
                           'coupon_discount': coupon_discount,
                           'additional_discount': additional_discount,
                           'delivery_charge_discount': delivery_charge_discount,
                           'saved_amount': invoice.discount_amount,
                           'note': order_instance.note if order_instance.note else None,
                           'colspan_value': "4" if is_offer else "3"
                           }

                admin_email = config("ORDER_NOTIFICATION_STAFF_EMAILS").replace(" ", "").split(',')
                html_admin = get_template('email/order_notification_staff.html')
                html_content = html_admin.render(content)
                msg_to_admin = EmailMultiAlternatives(admin_subject, 'shodai', from_email, admin_email)
                msg_to_admin.attach_alternative(html_content, "text/html")
                msg_to_admin.send()
            return Response({
                "status": "success",
                "message": "request received."
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "status": "failed",
                "message": "Invalid request!"
            }, status=status.HTTP_400_BAD_REQUEST)


class DeliveryChargeOfferList(APIView):
    permission_classes = [IsAdminUser]
    """
    Get All Offers with offer_types DD
    """

    def get(self, request):
        today = timezone.now()
        offers = Offer.objects.filter(offer_types="DD", is_approved=True, offer_starts_in__lte=today,
                                      offer_ends_in__gte=today)
        serializer = DeliveryChargeOfferSerializer(offers, many=True)
        return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)


class VerifyCoupon(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        data = request.data
        required_fields = ['coupon_code',
                           'order_id',
                           'products',
                           'mobile_number']
        is_valid = field_validation(required_fields, data)
        products = data['products']

        if is_valid and isinstance(products, list) and products:
            required_fields = ['product_id', 'product_quantity']
            product_list = []
            for item in products:
                is_valid = field_validation(required_fields, item)
                if is_valid and isinstance(item['product_id'], int):
                    if item['product_id'] not in product_list:
                        product_list.append(item['product_id'])
                        product_exist = Product.objects.filter(id=item['product_id'])
                        if not product_exist:
                            is_valid = False
                        if is_valid:
                            decimal_allowed = product_exist[0].decimal_allowed
                            if not decimal_allowed and not isinstance(item['product_quantity'], int) and \
                                    not item['product_quantity'] > 0:
                                is_valid = False
                            elif decimal_allowed and not isinstance(item['product_quantity'], (float, int)) and \
                                    not item['product_quantity'] > 0:
                                is_valid = False
                        if is_valid and decimal_allowed:
                            item['product_quantity'] = math.floor(item['product_quantity'] * 10 ** 3) / 10 ** 3
                    else:
                        is_valid = False
                else:
                    is_valid = False
                if not is_valid:
                    break
        else:
            is_valid = False

        if is_valid and len(data["mobile_number"]) == 14 and \
                data["mobile_number"].startswith('+8801') and data["mobile_number"][1:].isdigit():
            user = UserProfile.objects.filter(mobile_number=data["mobile_number"])
            if not user:
                is_valid = False
        else:
            is_valid = False

        if is_valid and data['order_id'] and isinstance(data['order_id'], int):
            invoice = InvoiceInfo.objects.filter(order_number=data['order_id']).order_by('-created_on')
            if invoice:
                discount = DiscountInfo.objects.filter(discount_type='CP', invoice=invoice[0])
                if discount and discount[0].coupon:
                    coupon_code = discount[0].coupon.coupon_code
                    is_used = True
                else:
                    is_valid = False
            else:
                is_valid = False
        elif is_valid and data['coupon_code'] and isinstance(data['coupon_code'], str):
            coupon_code = data['coupon_code']
            is_used = False
        else:
            is_valid = False

        if not is_valid:
            return Response({
                "status": "failed",
                "message": "Invalid request!"}, status=status.HTTP_400_BAD_REQUEST)

        discount_amount, coupon, _, is_under_limit = coupon_checker(coupon_code, products, user[0], is_used=is_used)
        if not is_under_limit:
            if discount_amount:
                return Response({'status': 'success',
                                 "discount_amount": discount_amount,
                                 "coupon_code": coupon.coupon_code}, status=status.HTTP_200_OK)
            elif discount_amount == 0:
                msg = "Discount is not applicable on products with offer."
                return Response({'status': 'failed',
                                 'message': msg}, status=status.HTTP_200_OK)
            return Response({'status': 'failed',
                             'message': "Invalid coupon!"}, status=status.HTTP_200_OK)
        else:
            msg = "Total price must be {} or more.".format(coupon.minimum_purchase_limit)
            return Response({'status': 'failed',
                             'message': msg}, status=status.HTTP_200_OK)


class UserList(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        search = request.query_params.get('search')
        sort_type = request.query_params.get('sort_type', 'dsc')
        sort_by = 'created_on'
        if sort_type != 'asc':
            sort_by = '-' + sort_by

        if search:
            if search.startswith("01") and len(search) == 11:
                queryset = UserProfile.objects.filter(mobile_number='+88' + search).order_by(sort_by)
            elif '@' in search:
                queryset = UserProfile.objects.filter(email=search).order_by(sort_by)
            else:
                queryset = UserProfile.objects.filter(Q(first_name__icontains=search) |
                                                      Q(last_name__icontains=search)).order_by(sort_by)
        else:
            queryset = UserProfile.objects.all().order_by(sort_by)
        paginator = CustomPageNumberPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = UserProfileSerializer(result_page, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)


class UserDetail(APIView):
    permission_classes = [IsAdminUser]
    """
    Get User Details
    """

    def get(self, request, id):
        user = get_object_or_404(UserProfile, id=id)
        serializer = UserProfileSerializer(user)
        return Response(serializer.data, status=status.HTTP_200_OK)


class UserListDownloadExcel(APIView):
    permission_classes = [IsAdminUserQP]
    """
    Get Excel of Users
    """

    def get(self, request):
        queryset = UserProfile.objects.filter(user_type='CM', is_active=True,
                                              is_approved=True).order_by('-created_on')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Customer List'

        columns = ["No.", "Name", "Mobile Number", "Email", "Registered on", "Order Count", "Last Order Placed"]
        sheet.append(columns)

        for count, customer in enumerate(queryset, 1):
            if customer.first_name and customer.last_name:
                full_name = customer.first_name + " " + customer.last_name
            elif customer.first_name:
                full_name = customer.first_name
            else:
                full_name = ""

            email = customer.email if customer.email else ""
            orders = Order.objects.filter(user=customer).order_by('-created_on').exclude(order_status='CN')
            order_count = orders.count()
            last_order_date = str(orders[0].placed_on + timedelta(hours=6))[:19] if orders else ""

            row = [count, full_name, str(customer.mobile_number[3:]), email,
                   str(customer.created_on + timedelta(hours=6))[:19],
                   order_count, last_order_date]
            sheet.append(row)

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=customer_list.xlsx'
        workbook.save(response)
        return response


class UserResetPassword(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        mobile_number = request.data.get('mobile_number')
        if mobile_number:
            user_instance = UserProfile.objects.filter(mobile_number=mobile_number)

        if not mobile_number or not user_instance:
            return Response({
                "status": "failed",
                "message": "Invalid request!"
            }, status=status.HTTP_400_BAD_REQUEST)

        user_instance = user_instance[0]
        temp_password = get_random_string(length=6)
        if not settings.DEBUG:
            sms_body = "Dear Customer,\n" + \
                       "We have created a new password [{}] ".format(temp_password) + \
                       "based on your reset password request. " + \
                       "Please change your password after login.\n\n" + \
                       "www.shod.ai"
            sms_flag = send_sms(user_instance.mobile_number, sms_body)
            if sms_flag == 'success':
                user_instance.set_password(temp_password)
                user_instance.save()
                return Response({'status': 'success',
                                 'message': "Password reset successful."}, status=status.HTTP_200_OK)
            else:
                return Response({'status': 'failed',
                                 'message': "Password reset failed."},
                                status=status.HTTP_200_OK)
        else:
            return Response({'status': 'failed',
                             'message': "Service not available."},
                            status=status.HTTP_200_OK)


class OrderProductListExcel(APIView):
    permission_classes = [IsAdminUserQP]

    def get(self, request):
        date_type = request.query_params.get('date_type', 'placed_on')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        product_meta = request.query_params.get('product_subcategory')
        order_status = all_order_status.get(request.query_params.get('order_status'))
        sort_by = 'created_on'

        try:
            date_from = timezone.make_aware(datetime.strptime(date_from, "%Y-%m-%d"))
        except Exception:
            if date_type == 'placed_on':
                date_from = Order.objects.first().placed_on
            else:
                date_from = Order.objects.first().delivery_date_time
        try:
            date_to = timezone.make_aware(datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except Exception:
            date_to = timezone.now()

        if date_type == 'placed_on':
            if product_meta and order_status:
                queryset = OrderProduct.objects.filter(product__product_category__type_of_product=product_meta,
                                                       order__order_status=order_status,
                                                       order__placed_on__gte=date_from,
                                                       order__placed_on__lt=date_to).order_by(sort_by)

            elif product_meta and not order_status:
                queryset = OrderProduct.objects.filter(product__product_category__type_of_product=product_meta,
                                                       order__placed_on__gte=date_from,
                                                       order__placed_on__lt=date_to).order_by(sort_by)

            elif not product_meta and order_status:
                queryset = OrderProduct.objects.filter(order__order_status=order_status,
                                                       order__placed_on__gte=date_from,
                                                       order__placed_on__lt=date_to).order_by(sort_by)
            else:
                queryset = OrderProduct.objects.filter(order__placed_on__gte=date_from,
                                                       order__placed_on__lt=date_to).order_by(sort_by)
        else:
            if product_meta and order_status:
                queryset = OrderProduct.objects.filter(product__product_category__type_of_product=product_meta,
                                                       order__order_status=order_status,
                                                       order__delivery_date_time__gte=date_from,
                                                       order__delivery_date_time__lt=date_to).order_by(sort_by)

            elif product_meta and not order_status:
                queryset = OrderProduct.objects.filter(product__product_category__type_of_product=product_meta,
                                                       order__delivery_date_time__gte=date_from,
                                                       order__delivery_date_time__lt=date_to).order_by(sort_by)

            elif not product_meta and order_status:
                queryset = OrderProduct.objects.filter(order__order_status=order_status,
                                                       order__delivery_date_time__gte=date_from,
                                                       order__delivery_date_time__lt=date_to).order_by(sort_by)
            else:
                queryset = OrderProduct.objects.filter(order__delivery_date_time__gte=date_from,
                                                       order__delivery_date_time__lt=date_to).order_by(sort_by)

        if queryset:
            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to')
            product_meta = request.query_params.get('product_subcategory', 'All')
            order_status = request.query_params.get('order_status', 'All')

            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'Order Product List'

            field_names = ["No.", "Customer Mobile Number", "Order Number", "Order Placing Date Time ",
                           "Order Delivery Date Time", "Order Status", "Order Total Amount", "Product ID",
                           "Product Name", "Product Unit", "Product Unit Price", "Product Quantity",
                           "Product Total Price", "Product Subcategory", "Delivery Address"]
            ws1.append(field_names)

            for count, obj in enumerate(queryset, 1):
                row = [count, obj.order.user.mobile_number[3:], obj.order.order_number,
                       str(obj.order.placed_on + timedelta(hours=6))[:16],
                       str(obj.order.delivery_date_time + timedelta(hours=6))[:16],
                       order_status_all[obj.order.order_status], obj.order.order_total_price, obj.product.id,
                       obj.product.product_name, obj.product.product_unit.product_unit, obj.order_product_price,
                       obj.order_product_qty, obj.order_product_price * obj.order_product_qty,
                       obj.product.product_category.type_of_product, obj.order.address.road]
                ws1.append(row)

            for column_cells in ws1.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws1.column_dimensions[column_cells[0].column_letter].width = length + 1

            ws1.insert_rows(idx=1, amount=7)
            report_info = [f"Title: Order Product List",
                           f"Date Type: {'Placing Date' if date_type == 'placed_on' else 'Delivery Date'}",
                           f"Date Range: {date_from} to {date_to}",
                           f"Order Status: {order_status}",
                           f"Product Subcategory: {product_meta}",
                           f"Generated on {str(timezone.now() + timedelta(hours=6))[:19]}",
                           f""]
            for count, line in enumerate(report_info, 1):
                ws1["A" + str(count)] = line
                ws1.merge_cells(start_row=count, start_column=1, end_row=count, end_column=ws1.max_column)

            order_product_list = []
            count = 0
            for obj in queryset:
                for item in order_product_list:
                    if item['Product Name'] == obj.product.product_name:
                        item['Product Quantity'] += obj.order_product_qty
                        break
                else:
                    count += 1
                    product_data = {'No.': count,
                                    'Product ID': obj.product.id,
                                    'Product Name': obj.product.product_name,
                                    'Product Unit': obj.product.product_unit.product_unit,
                                    'Product Quantity': obj.order_product_qty,
                                    'Product Subcategory': obj.product.product_category.type_of_product}
                    order_product_list.append(product_data)

            ws2 = wb.create_sheet('Order Product List (Summary)')
            ws2.append(list(order_product_list[0].keys()))
            for count, product in enumerate(order_product_list):
                row = list(product.values())
                ws2.append(row)

            for column_cells in ws2.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws2.column_dimensions[column_cells[0].column_letter].width = length + 1

            ws2.insert_rows(idx=1, amount=7)
            report_info = [f"Title: Order Product List (Summary)",
                           f"Date Type: {'Placing Date' if date_type == 'placed_on' else 'Delivery Date'}",
                           f"Date Range: {date_from} to {date_to}",
                           f"Order Status: {order_status}",
                           f"Product Subcategory: {product_meta}",
                           f"Generated on {str(timezone.now() + timedelta(hours=6))[:19]}",
                           f""]
            for count, line in enumerate(report_info, 1):
                ws2["A" + str(count)] = line
                ws2.merge_cells(start_row=count, start_column=1, end_row=count, end_column=ws2.max_column)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=order_product_list.xlsx'
            wb.save(response)
            return response
        else:
            template = get_template("others/report_no_data.html")
            return HttpResponse(template.render())


class ProductMetaList(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        queryset = ProductMeta.objects.filter(is_approved=True)
        serializer = ProductMetaSerializer(queryset, many=True)
        return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)


class OrderStatusUpdate(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        data = request.data

        order_list = data['order_list']
        order_status = all_order_status.get(data['order_status'])
        if order_status:
            for order_id in order_list:
                order = Order.objects.filter(id=order_id)
                if order and order[0].order_status != 'COM' and order[0].order_status != 'CN':
                    order[0].order_status = order_status
                    order[0].save()
            return Response({'status': 'success',
                             'message': "Order status updated."}, status=status.HTTP_200_OK)
        else:
            return Response({"status": "failed",
                             "message": "Invalid request!"}, status=status.HTTP_400_BAD_REQUEST)


class PreOrderSettingList(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        queryset = PreOrderSetting.objects.all().order_by('-created_on')
        paginator = CustomPageNumberPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = PreOrderSettingListSerializer(result_page, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)

    def post(self, request):
        data = request.data
        required_fields = ['producer_product_id',
                           'product_id',
                           'start_date',
                           'end_date',
                           'delivery_date',
                           'discounted_price',
                           'unit_quantity',
                           'target_quantity']

        is_valid = field_validation(required_fields, data)
        if is_valid:
            string_fields = [data['start_date'],
                             data['end_date'],
                             data['delivery_date']]
            is_valid = type_validation(string_fields, str)
        if is_valid:
            number_fields = [data['producer_product_id'],
                             data['product_id'],
                             data['discounted_price'],
                             data['unit_quantity'],
                             data['target_quantity']]
            is_valid = type_validation(number_fields, (float, int))

        if is_valid and data['start_date']:
            try:
                start_date = timezone.make_aware(datetime.strptime(data['start_date'], "%Y-%m-%d"))
            except Exception:
                is_valid = False
        if is_valid and data['end_date']:
            try:
                end_date = timezone.make_aware(datetime.strptime(data['end_date'], "%Y-%m-%d").replace(hour=23, minute=59, second=59))
            except Exception:
                is_valid = False
        if is_valid and data['delivery_date']:
            try:
                delivery_date = timezone.make_aware(datetime.strptime(data['delivery_date'], "%Y-%m-%d"))
            except Exception:
                is_valid = False
        if is_valid:
            product = Product.objects.filter(id=data['product_id'], is_approved=True).first()
            producer_product = ProducerProductRequest.objects.filter(id=data['producer_product_id'],
                                                                     is_approved=True).first()
            if not product or not producer_product or not start_date < end_date < delivery_date or \
                    data['discounted_price'] > product.product_price or \
                    data['unit_quantity'] > data['target_quantity'] or \
                    data['target_quantity'] % data['unit_quantity']:
                is_valid = False
        if not is_valid:
            return Response({
                "status": "failed",
                "message": "Invalid request!"}, status=status.HTTP_400_BAD_REQUEST)

        pre_order_exists = PreOrderSetting.objects.filter(producer_product=producer_product).exists()
        if pre_order_exists:
            return Response({
                "status": "failed",
                "message": "Pre-order setting already exist!"}, status=status.HTTP_200_OK)

        pre_order_setting = PreOrderSetting.objects.create(pre_order_setting_number=str(uuid.uuid4())[:4].upper(),
                                                           producer_product=producer_product,
                                                           product=product,
                                                           start_date=start_date,
                                                           end_date=end_date,
                                                           delivery_date=delivery_date,
                                                           discounted_price=data['discounted_price'],
                                                           unit_quantity=data['unit_quantity'],
                                                           target_quantity=data['target_quantity'])
        return Response({"status": "success",
                         "message": "Pre-order setting created.",
                         "pre_order_setting_id": pre_order_setting.id}, status=status.HTTP_201_CREATED)


class PreOrderSettingDetail(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, id):
        pre_order_setting = get_object_or_404(PreOrderSetting, id=id)
        serializer = PreOrderSettingDetailSerializer(pre_order_setting)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, id):
        data = request.data
        required_fields = ['product_id',
                           'start_date',
                           'end_date',
                           'delivery_date',
                           'discounted_price',
                           'unit_quantity',
                           'target_quantity',
                           'is_approved']

        is_valid = field_validation(required_fields, data)
        if is_valid:
            string_fields = [data['start_date'],
                             data['end_date'],
                             data['delivery_date']]
            is_valid = type_validation(string_fields, str)
        if is_valid:
            number_fields = [data['product_id'],
                             data['discounted_price'],
                             data['unit_quantity'],
                             data['target_quantity']]
            is_valid = type_validation(number_fields, (float, int))

        if is_valid and data['start_date']:
            try:
                start_date = timezone.make_aware(datetime.strptime(data['start_date'], "%Y-%m-%d"))
            except Exception:
                is_valid = False
        if is_valid and data['end_date']:
            try:
                end_date = timezone.make_aware(datetime.strptime(data['end_date'], "%Y-%m-%d").replace(hour=23, minute=59, second=59))
            except Exception:
                is_valid = False
        if is_valid and data['delivery_date']:
            try:
                delivery_date = timezone.make_aware(datetime.strptime(data['delivery_date'], "%Y-%m-%d"))
            except Exception:
                is_valid = False
        if is_valid:
            pre_order_setting = PreOrderSetting.objects.filter(id=id).first()
            product = Product.objects.filter(id=data['product_id'], is_approved=True).first()
            if not pre_order_setting or not product or not start_date < end_date < delivery_date or \
                    data['discounted_price'] > product.product_price or \
                    data['unit_quantity'] > data['target_quantity'] or \
                    data['target_quantity'] % data['unit_quantity'] or \
                    not isinstance(data['is_approved'], bool):
                is_valid = False
        if not is_valid:
            return Response({
                "status": "failed",
                "message": "Invalid request!"}, status=status.HTTP_400_BAD_REQUEST)

        if pre_order_setting.end_date > timezone.now():
            pre_order_setting.product = product
            pre_order_setting.start_date = start_date
            pre_order_setting.end_date = end_date
            pre_order_setting.unit_quantity = data['unit_quantity']
            pre_order_setting.target_quantity = data['target_quantity']
            pre_order_setting.is_approved = data['is_approved']
        pre_order_setting.delivery_date = delivery_date
        pre_order_setting.discounted_price = data['discounted_price']
        pre_order_setting.modified_by = request.user
        pre_order_setting.save()

        return Response({"status": "success",
                         "message": "Pre-order setting updated.",
                         "pre_order_setting_id": pre_order_setting.id}, status=status.HTTP_200_OK)


class ProducerProductSearch(APIView):
    permission_classes = [IsAdminUser]
    """
    Get List of Producer Product by Producer mobile number
    """

    def get(self, request):
        query = request.query_params.get('mobile_number', '')
        producer = UserProfile.objects.filter(mobile_number='+88' + query,
                                              is_producer=True).first()
        if producer:
            queryset = ProducerProductRequest.objects.filter(producer=producer, is_approved=True)
            serializer = ProducerProductSerializer(queryset, many=True)
            return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)
        else:
            return Response({'status': 'failed',
                             'message': 'Producer not found.'}, status=status.HTTP_200_OK)


class ProcessPreOrder(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        data = request.data
        user = request.user
        required_fields = ['pre_order_setting_id',
                           'action']

        is_valid = field_validation(required_fields, data)

        if is_valid and isinstance(data['pre_order_setting_id'], int):
            pre_order_setting = PreOrderSetting.objects.filter(id=data['pre_order_setting_id'],
                                                               is_approved=True,
                                                               is_processed=False).first()

            if not pre_order_setting:
                is_valid = False
        else:
            is_valid = False

        if not is_valid or not isinstance(data['action'], str) or not data['action'] in ['place', 'cancel']:
            return Response({
                "status": "failed",
                "message": "Invalid request!"}, status=status.HTTP_400_BAD_REQUEST)

        pre_orders = PreOrder.objects.filter(pre_order_setting=pre_order_setting).exclude(pre_order_status='CN')
        if not pre_orders:
            return Response({
                "status": "failed",
                "message": "No pre-order against this setting!"}, status=status.HTTP_200_OK)

        if data['action'] == 'place':
            for p in pre_orders:
                order = Order.objects.create(platform=p.platform,
                                             delivery_date_time=pre_order_setting.delivery_date,
                                             delivery_place='Dhaka',
                                             lat=23.777176,
                                             long=90.399452,
                                             contact_number=p.contact_number,
                                             address=p.delivery_address,
                                             note=p.note,
                                             user=p.customer,
                                             created_by=user)
                order_product = OrderProduct.objects.create(product=pre_order_setting.product,
                                                            product_price=pre_order_setting.product.product_price,
                                                            order=order,
                                                            order_product_price=pre_order_setting.discounted_price,
                                                            order_product_qty=p.product_quantity)

                sub_total = float(order_product.order_product_price) * order_product.order_product_qty
                sub_total_without_discount = float(order_product.product_price) * order_product.order_product_qty
                total_vat = float(
                    order_product.order_product_price_with_vat - order_product.order_product_price) * order_product.order_product_qty
                delivery_charge = DeliveryCharge.objects.get().delivery_charge_inside_dhaka

                order.total_vat = total_vat
                order.order_total_price = sub_total + total_vat + delivery_charge
                order.save()

                if p.customer.first_name and p.customer.last_name:
                    billing_person_name = p.customer.first_name + " " + p.customer.last_name
                elif p.customer.first_name:
                    billing_person_name = p.customer.first_name
                else:
                    billing_person_name = ""

                invoice = InvoiceInfo.objects.create(invoice_number=order.invoice_number,
                                                     billing_person_name=billing_person_name,
                                                     billing_person_email=p.customer.email,
                                                     billing_person_mobile_number=p.customer.mobile_number,
                                                     delivery_contact_number=order.contact_number,
                                                     delivery_address=order.address.road,
                                                     delivery_date_time=order.delivery_date_time,
                                                     delivery_charge=delivery_charge,
                                                     discount_amount=sub_total_without_discount - sub_total,
                                                     net_payable_amount=order.order_total_price,
                                                     payment_method='CASH_ON_DELIVERY',
                                                     order_number=order,
                                                     user=p.customer,
                                                     created_by=user)
                DiscountInfo.objects.create(discount_amount=sub_total_without_discount - sub_total,
                                            discount_type='PD',
                                            discount_description='Product Offer Discount',
                                            invoice=invoice)
                p.order = order
                p.pre_order_status = 'OA'
                p.save()

                if not settings.DEBUG:
                    async_task('order.tasks.send_order_email', order, True)

            pre_order_setting.is_processed = True
            pre_order_setting.modified_by = user
            pre_order_setting.save()

            return Response({'status': 'success',
                             'message': 'Orders created.'}, status=status.HTTP_200_OK)
        else:
            for p in pre_orders:
                p.pre_order_status = 'CN'
                p.save()

            pre_order_setting.is_processed = True
            pre_order_setting.modified_by = user
            pre_order_setting.save()
            return Response({'status': 'success',
                             'message': 'Pre-orders cancelled.'}, status=status.HTTP_200_OK)


# class PreOrderSettingSearch(APIView):
#     permission_classes = [IsAdminUser]
#
#     def get(self, request):
#         product_name = request.query_params.get('query', '')
#         queryset = PreOrderSetting.objects.filter(product__product_name__icontains=product_name,
#                                                   is_approved=True).order_by('-created_on')
#         serializer = PreOrderSettingListSerializer(queryset, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)


class PreOrderStatusList(APIView):
    permission_classes = [IsAdminUser]
    """
    Get All Pre-Order Status
    """

    def get(self, request):
        pre_order_status = [
            'Ordered',
            'Order Accepted',
            'Order Cancelled',
        ]
        return Response({'status': 'success',
                         'data': pre_order_status}, status=status.HTTP_200_OK)


class PreOrderList(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        search = request.query_params.get('search')
        pre_order_status = all_order_status.get(request.query_params.get('pre_order_status'))

        if search and pre_order_status:
            if search.startswith("01") and len(search) == 11:
                queryset = PreOrder.objects.filter(customer__mobile_number='+88'+search,
                                                   pre_order_status=pre_order_status).order_by('-created_on')
            else:
                queryset = PreOrder.objects.filter(pre_order_setting__id=search,
                                                   pre_order_status=pre_order_status).order_by('-created_on')
        elif search and not pre_order_status:
            if search.startswith("01") and len(search) == 11:
                queryset = PreOrder.objects.filter(customer__mobile_number='+88'+search).order_by('-created_on')
            else:
                queryset = PreOrder.objects.filter(pre_order_setting__id=search).order_by('-created_on')
        elif not search and pre_order_status:
            queryset = PreOrder.objects.filter(pre_order_status=pre_order_status).order_by('-created_on')
        else:
            queryset = PreOrder.objects.all().order_by('-created_on')

        paginator = CustomPageNumberPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = PreOrderListSerializer(result_page, many=True, context={'request': request})
        return paginator.get_paginated_response(serializer.data)


class PreOrderDetail(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, id):
        pre_order = get_object_or_404(PreOrder, id=id)
        serializer = PreOrderDetailSerializer(pre_order)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, id):
        data = request.data
        required_fields = ['delivery_address',
                           'contact_number',
                           'product_quantity',
                           'note',
                           'pre_order_status']

        is_valid = field_validation(required_fields, data)
        if is_valid:
            string_fields = [data['delivery_address'],
                             data['contact_number'],
                             data['note'],
                             data['pre_order_status']]
            is_valid = type_validation(string_fields, str)

        if is_valid and len(data["contact_number"]) == 14 and \
                data["contact_number"].startswith('+8801') and data["contact_number"][1:].isdigit():
            pass
        else:
            is_valid = False

        if is_valid:
            pre_order = PreOrder.objects.filter(id=id, pre_order_setting__is_processed=False).first()
            if not pre_order:
                is_valid = False

        if not is_valid or not isinstance(data['product_quantity'], int) or not data['product_quantity'] > 0 or \
                not data['delivery_address'] or data['pre_order_status'] not in all_order_status:
            return Response({
                "status": "failed",
                "message": "Invalid request!"}, status=status.HTTP_400_BAD_REQUEST)

        data["delivery_address"] = data["delivery_address"][:500]
        if data["delivery_address"] != pre_order.delivery_address.road:
            delivery_address = Address.objects.filter(road=data["delivery_address"]).first()
            if not delivery_address:
                delivery_address = Address.objects.create(road=data["delivery_address"],
                                                          city="Dhaka",
                                                          district="Dhaka",
                                                          country="Bangladesh",
                                                          zip_code="",
                                                          user=pre_order.customer)
        else:
            delivery_address = pre_order.delivery_address

        pre_order.delivery_address = delivery_address
        pre_order.contact_number = data["contact_number"]
        pre_order.product_quantity = data['product_quantity']
        pre_order.note = data['note'][:500]
        pre_order.pre_order_status = all_order_status[data['pre_order_status']]
        pre_order.modified_by = request.user
        pre_order.save()

        return Response({"status": "success",
                         "message": "Pre-order updated.",
                         "pre_order_id": pre_order.id}, status=status.HTTP_200_OK)


class DeliveryZoneList(APIView):
    permission_classes = [IsAdminUser]
    """
    Get List of Delivery zones
    """

    def get(self, request):
        queryset = DeliveryZone.objects.filter(is_approved=True)
        serializer = DeliveryZoneSerializer(queryset, many=True)
        return Response({'status': 'success', 'data': serializer.data}, status=status.HTTP_200_OK)
